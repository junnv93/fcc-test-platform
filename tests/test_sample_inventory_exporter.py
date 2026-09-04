"""PM/RF 엑셀 내보내기 — 렌더 경로가 실제로 지나간다 (2026-09-04).

⚠️ 이 파일이 생긴 이유는 **이 경로를 지나가는 테스트가 한 건도 없었기 때문**이다.
결함이 둘 겹쳐 있었고 어느 것도 드러나지 않았다:

  ① 디렉터리 이름이 어긋나 있었다(`sample_inventory_templates` vs 실제
     `sample_inventory`) → 개발 트리에서도 모든 내보내기가 FileNotFoundError.
  ② 이름을 고쳐도 **컨테이너에서는 여전히 못 찾는다.** 렌더러가
     `parents[3] / 'tests' / 'fixtures'` 로 패키지 밖으로 걸어 나갔는데, 이미지는
     `tests/` 를 COPY 하지 않고 휠은 `fcc_test_platform*` 만 싣는다.

②는 ①을 고친 사람(=이 파일을 처음 쓴 사람)이 못 본 축이다. 「고쳤다」와 「돌고 있는
것이 고쳐졌다」가 다른 축이라는 것을 동료 세션이 짚어 주었다. 그래서 이 파일은
**해소 · 렌더 · 설치 형태** 셋을 각각 잡는다 — 마지막 것이 없으면 ②가 다시 들어온다.
"""
from __future__ import annotations

import importlib
import io
from importlib.resources import as_file
from pathlib import Path
import tomllib

import openpyxl
import pytest

from fcc_test_platform.infrastructure.excel.sample_inventory_exporter import (
    PM_HEADERS,
    TEMPLATE_FILES,
    TEMPLATE_PACKAGE,
    TEMPLATE_SUBDIRECTORY,
    SampleInventoryExcelExporter,
    template_resource,
)


PROJECT = {'model_name': 'SM-F968U1', 'project_code': 'P-1'}


def _pm_rows(content: bytes) -> list[list]:
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(min_row=4, values_only=True)]


class TestTheTemplatesAreWhereTheCodeLooks:
    """경로 오타가 다시 들어오면 여기서 멈춘다."""

    @pytest.mark.parametrize('template', sorted(TEMPLATE_FILES))
    def test_every_declared_template_resolves(self, template):
        resource = template_resource(template)
        assert resource.is_file(), (
            f'{template} 템플릿이 {resource} 에 없다 — 내보내기가 런타임에 죽는다.'
        )

    def test_pm_and_rf_render_to_a_readable_workbook(self):
        exporter = SampleInventoryExcelExporter()
        for template in sorted(TEMPLATE_FILES):
            content = exporter.render(template, [], project=PROJECT)
            assert content[:2] == b'PK', f'{template} 이 xlsx 바이트를 내지 않았다'


class TestTheTemplatesShipInsideThePackage:
    """②의 자리 — 「고쳤다」가 아니라 「배포된 것이 고쳐졌다」를 잰다.

    컨테이너는 `pip install --no-deps .` 뒤 소스를 지우고 site-packages 만으로 돈다.
    템플릿이 패키지 밖에 있으면 개발 트리에서는 모든 시험이 초록인 채로 운영에서만
    죽는다 — 이 저장소가 `decision_catalogue.json` 으로 이미 한 번 치른 값이다.
    """

    def test_the_resource_lives_under_the_exporter_package(self):
        assert TEMPLATE_PACKAGE == 'fcc_test_platform.infrastructure.excel'

    @pytest.mark.parametrize('template', sorted(TEMPLATE_FILES))
    def test_the_resource_is_not_reached_by_walking_out_of_the_package(self, template):
        """패키지 경계 밖으로 나가면 배포에서 못 따라온다."""
        # ⚠️ `__file__` 이 아니라 `__path__` 다 — 이 배포판의 최상위는 PEP 420
        # 네임스페이스 패키지라 `__file__` 이 `None` 이고, 잎 패키지가 언제 그렇게
        # 될지는 이 시험이 결정하지 않는다.
        package_root = Path(
            list(importlib.import_module(TEMPLATE_PACKAGE).__path__)[0]
        ).resolve()
        with as_file(template_resource(template)) as path:
            assert path.resolve().is_relative_to(package_root), (
                f'{template} 이 패키지 밖({path})에서 해소된다 — 휠에는 실리지 않는다.'
            )

    def test_the_declaration_covers_every_template(self):
        """pyproject 의 package-data 선언이 파일 집합과 어긋나면 휠에서 빠진다."""
        declaration = tomllib.loads(
            (Path(__file__).resolve().parent.parent / 'pyproject.toml')
            .read_text(encoding='utf-8')
        )['tool']['setuptools']['package-data']
        patterns = declaration.get(TEMPLATE_PACKAGE, [])
        assert f'{TEMPLATE_SUBDIRECTORY}/*.xlsx' in patterns, (
            f'package-data 가 {TEMPLATE_PACKAGE} 의 템플릿을 선언하지 않는다: {patterns}'
        )


class TestClassificationReachesTheWorkbook:
    def test_sample_kind_is_read_from_the_row_not_hardcoded(self):
        """ADR-0002 결정 8 이전에는 이 칸이 'Device' 리터럴이었다."""
        records = [
            {'sample_number': '#1', 'sample_kind': 'Device',
             'sample_description': 'SM-F968U1_Main Conduction #1_Dummy Batt',
             'test_category': 'Conduction'},
            {'sample_number': '#2', 'sample_kind': 'Accessory',
             'sample_description': 'SM-F968U1_Dummy Batt', 'test_category': None},
        ]
        rows = _pm_rows(SampleInventoryExcelExporter().render(
            'pm-status', records, project=PROJECT))
        kind_column = PM_HEADERS.index('Sample')
        assert [row[kind_column] for row in rows] == ['Device', 'Accessory']

    def test_the_stored_description_is_used_verbatim(self):
        """파생식 '{model}_{category} {number}' 로는 만들 수 없는 값이 살아남는다."""
        records = [{'sample_number': '#1', 'sample_kind': 'Device',
                    'sample_description': 'SM-F968U1_Main Conduction #1_Dummy Batt',
                    'test_category': 'Conduction'}]
        rows = _pm_rows(SampleInventoryExcelExporter().render(
            'pm-status', records, project=PROJECT))
        description_column = PM_HEADERS.index('Sample Description')
        assert rows[0][description_column] == 'SM-F968U1_Main Conduction #1_Dummy Batt'

    def test_an_unfilled_description_falls_back_to_the_previous_derivation(self):
        """새 컬럼이 비어 있다고 해서 이미 나가던 엑셀이 비어 보이면 안 된다."""
        records = [{'sample_number': '#1', 'test_category': 'Conduction'}]
        rows = _pm_rows(SampleInventoryExcelExporter().render(
            'pm-status', records, project=PROJECT))
        description_column = PM_HEADERS.index('Sample Description')
        kind_column = PM_HEADERS.index('Sample')
        assert rows[0][description_column] == 'SM-F968U1_Conduction #1'
        assert rows[0][kind_column] == 'Device'


class TestTheRfSheetsStillSplitOnTestCategory:
    def test_conduction_and_radiation_rows_land_on_their_own_sheets(self):
        records = [
            {'sample_number': '#1', 'test_category': 'Conduction',
             'intake': {'bl': 'BL-C'}},
            {'sample_number': '#2', 'test_category': 'Radiation',
             'intake': {'bl': 'BL-R'}, 'radiation_tech_group': 'G1'},
        ]
        content = SampleInventoryExcelExporter().render(
            'rf-data', records, project=PROJECT)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        for sheet_name, expected_bl in (('Conduction', 'BL-C'), ('Radiation', 'BL-R')):
            rows = [list(row) for row in
                    workbook[sheet_name].iter_rows(min_row=3, values_only=True)]
            assert [row[1] for row in rows] == [expected_bl], sheet_name
