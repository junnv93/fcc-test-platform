"""openpyxl renderer for the sanitized PM and RF sample templates."""
from __future__ import annotations

from contextlib import ExitStack
from copy import copy
from datetime import datetime
from importlib.resources import as_file as _resource_as_file
from importlib.resources import files as _resource_files
from pathlib import Path
import tempfile
from typing import Any, Mapping, Sequence

import openpyxl

from fcc_test_kernel.infrastructure.excel.atomic_write import atomic_xlsx_write


PM_HEADERS = (
    '', 'Model name', 'Sample', 'Sample Description', '라벨넘버', 'SMSN',
    'S/N or\nIMEI', '반입증', 'TEAM', '발신자/연락처', '수신자/연락처',
    '수령한 날짜', '반출한 날짜', 'Note', '',
)
RF_HEADERS = (
    '시료번호', 'BL', 'AP', 'CP', 'CSC', 'RF CAL', 'HW Rev', '비고', '입고일',
)
RF_KEYSTRING = '*#12580*369#'
XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
# ⚠️ **템플릿은 패키지 데이터다 — 저장소 상대 경로가 아니다.** 2026-09-04 이전에는
# 이 자리가 `Path(__file__).resolve().parents[3] / 'tests' / 'fixtures' / …` 였고
# 결함이 둘 겹쳐 있었다:
#
#   ① 디렉터리 이름이 'sample_inventory_templates' 인데 파일은 'sample_inventory' 에
#      있었다 → 개발 트리에서도 모든 내보내기가 FileNotFoundError 로 죽었다.
#   ② 이름을 고쳐도 **컨테이너에서는 여전히 못 찾는다.** 이미지는
#      `pip install --no-deps .` 뒤 `rm -rf /app/fcc_test_platform` 하고 전부
#      site-packages 에서 돈다. 휠은 `fcc_test_platform*` 만 싣고 `tests/` 는
#      이미지에 COPY 되지도 않는다 — parents[3] 는 `site-packages/tests/…` 로
#      해소된다.
#
# ②가 이 저장소가 이미 이름 붙인 결함 계급이다(pyproject `package-data` 주석:
# *"패키지 데이터는 import 가 아니다 … 상자에는 실려 있고 휠에만 없었다"*), 그리고
# `tree_artifacts` 가 이름 붙인 계급이기도 하다 — *"조상 몇 칸 위인지를 부호화하면
# 그 트리 하나에서만 맞다"*. 그래서 위로 걸어 나가는 대신 **패키지 안에서** 읽는다.
# 이 자리는 소스 트리와 설치된 휠에서 같은 답을 준다.
#
# 봉인: tests/test_sample_inventory_exporter.py (해소 · 렌더 · 휠 적재 셋 다).
TEMPLATE_PACKAGE = __package__
TEMPLATE_SUBDIRECTORY = 'templates'
TEMPLATE_FILES = {
    'pm-status': 'pm_sample_status.xlsx',
    'rf-data': 'rf_sample_data.xlsx',
}


def template_resource(template: str):
    """Return the Traversable for one export template (source tree or wheel)."""
    return (
        _resource_files(TEMPLATE_PACKAGE)
        .joinpath(TEMPLATE_SUBDIRECTORY)
        .joinpath(TEMPLATE_FILES[template])
    )


class SampleInventoryExcelExporter:
    """Render a workbook to bytes after atomic write + ZIP validation."""

    def render(self, template: str, records: Sequence[Mapping[str, Any]], *,
               project: Mapping[str, Any]) -> bytes:
        with tempfile.TemporaryDirectory(prefix='fcc-sample-export-') as directory:
            target = Path(directory) / 'sample-inventory.xlsx'
            atomic_xlsx_write(
                str(target),
                lambda temporary: self._write(temporary, template, records, project),
                validate=True,
                label='SampleInventoryExcelExporter',
            )
            return target.read_bytes()

    def _write(self, temp_path: str, template: str, records: Sequence[Mapping[str, Any]],
               project: Mapping[str, Any]) -> None:
        if template == 'pm-status':
            _write_pm(temp_path, records, project)
        elif template == 'rf-data':
            _write_rf(temp_path, records, project)
        else:
            raise ValueError(f'unsupported sample inventory export template: {template}')


def _write_pm(temp_path: str, records: Sequence[Mapping[str, Any]], project: Mapping[str, Any]) -> None:
    model = str(project.get('model_name') or project.get('project_code') or 'Samples')
    workbook = _load_template('pm-status')
    worksheet = workbook.active
    worksheet.title = _sheet_name(model)
    _replace_model_in_title(worksheet, model)
    # Keep the operational A:O footprint even though the final column is a
    # deliberately blank template column and openpyxl otherwise drops it.
    worksheet.cell(row=2, column=15).value = ''
    styles, row_height = _clear_dynamic_rows(worksheet, start_row=4)
    for index, sample in enumerate(records):
        row = worksheet.max_row + 1
        worksheet.append([
            None,
            model if index == 0 else None,
            # ADR-0002 결정 8까지 이 칸은 'Device' 하드코딩이었다 — Device/Accessory 를
            # 저장하는 곳이 아예 없었다. 기존 행은 sample_kind 가 비어 있으므로
            # 예전과 같은 값으로 되돌아간다.
            sample.get('sample_kind') or 'Device',
            _sample_description(sample, model),
            sample.get('label_number'),
            sample.get('smsn'),
            sample.get('serial_number'),
            sample.get('intake_cert'),
            sample.get('assigned_team'),
            sample.get('sender'),
            sample.get('receiver'),
            _excel_date(sample.get('received_date')),
            _excel_date(sample.get('released_date')),
            sample.get('note'),
            None,
        ])
        _copy_row_style(worksheet, row, styles, row_height)
    workbook.save(temp_path)
    workbook.close()


def _sample_description(sample: Mapping[str, Any], model: str) -> str:
    """PM 이 적은 구분 이름 — 없으면 예전의 파생식으로 되돌아간다 (ADR-0002 결정 2).

    파생식 ``{model}_{test_category} {sample_number}`` 은 **손실적**이다: 실제 값
    'SM-TEST1_Main Conduction #1_Dummy Batt' 의 'Main' 과 '_Dummy Batt' 를 만들어낼 수
    없다. 그래서 저장 컬럼이 생겼다. 되돌아가는 가지를 남기는 이유는 이 변경이
    **기존 행의 엑셀 출력을 바꾸지 않게** 하기 위해서다 — 아직 아무도 채우지 않은
    컬럼 때문에 이미 나가던 파일이 비어 보이면 안 된다.
    """
    stored = (sample.get('sample_description') or '').strip()
    if stored:
        return stored
    category = sample.get('test_category') or ''
    sample_number = sample.get('sample_number') or ''
    return f'{model}_{category} {sample_number}'.strip()


def _write_rf(temp_path: str, records: Sequence[Mapping[str, Any]], project: Mapping[str, Any]) -> None:
    workbook = _load_template('rf-data')
    for sheet_name in ('Conduction', 'Radiation'):
        worksheet = workbook[sheet_name]
        styles, row_height = _clear_dynamic_rows(worksheet, start_row=3)
        for record in records:
            if record.get('test_category') != sheet_name:
                continue
            intake = dict(record.get('intake') or {})
            row = worksheet.max_row + 1
            worksheet.append([
                record.get('sample_number'), intake.get('bl'), intake.get('ap'),
                intake.get('cp'), intake.get('csc'), intake.get('rf_cal'),
                intake.get('hw_rev'), intake.get('note'),
                _excel_date(intake.get('intake_date')),
                record.get('radiation_tech_group') if sheet_name == 'Radiation' else None,
            ][:10 if sheet_name == 'Radiation' else 9])
            _copy_row_style(worksheet, row, styles, row_height)
    workbook.save(temp_path)
    workbook.close()


def _load_template(template: str):
    # openpyxl 은 실제 파일 경로를 요구하므로 `as_file` 로 잠깐 실체화한다. 설치된
    # 휠이 디렉터리로 풀려 있으면 그 자리 그대로이고, 압축된 배포에서는 임시 사본이
    # 만들어진다 — 부르는 쪽은 어느 쪽인지 알 필요가 없다.
    resource = template_resource(template)
    with ExitStack() as stack:
        try:
            path = stack.enter_context(_resource_as_file(resource))
        except (FileNotFoundError, OSError) as exc:
            raise FileNotFoundError(
                f'sample inventory export template is missing: {resource}'
            ) from exc
        if not path.is_file():
            raise FileNotFoundError(
                f'sample inventory export template is missing: {path}'
            )
        workbook = openpyxl.load_workbook(path, data_only=False, keep_links=True)
    # openpyxl normalizes empty inline strings to numeric empty cells on save.
    # Retain the template's empty-cell type so the structural replica remains
    # semantically stable across the render round trip.
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None and cell.data_type == 'inlineStr':
                    cell.value = ''
    return workbook


def _replace_model_in_title(worksheet, model: str) -> None:
    title = worksheet['B1'].value
    if isinstance(title, str) and title:
        worksheet['B1'] = title.replace('SM-TEST1', model)


def _clear_dynamic_rows(worksheet, *, start_row: int):
    styles = [copy(cell._style) for cell in worksheet[start_row]]
    row_dimension = worksheet.row_dimensions[start_row]
    row_height = row_dimension.height
    if worksheet.max_row >= start_row:
        worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)
    return styles, row_height


def _copy_row_style(worksheet, row: int, styles, row_height) -> None:
    if row_height is not None:
        worksheet.row_dimensions[row].height = row_height
    for column, style in enumerate(styles, start=1):
        worksheet.cell(row=row, column=column)._style = copy(style)


def _excel_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        try:
            return datetime.fromisoformat(text.replace('Z', '+00:00')).replace(tzinfo=None)
        except ValueError:
            return value
    return value


def _sheet_name(value: str) -> str:
    cleaned = ''.join('_' if char in '[]:*?/\\' else char for char in value)
    return (cleaned[:31] or 'Samples').strip("'") or 'Samples'


__all__ = [
    'PM_HEADERS',
    'RF_HEADERS',
    'RF_KEYSTRING',
    'SampleInventoryExcelExporter',
    'XLSX_CONTENT_TYPE',
]
